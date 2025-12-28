import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def extract_usd_data_with_headers(excel_file):
    """
    Extracts only USD Mn columns from NSDL FII Excel file
    Preserves all header rows including "AUC as on December 15, 2025"
    """
    
    print("=" * 70)
    print("NSDL FII Data - USD Extraction (With Headers)")
    print("=" * 70)
    
    try:
        # Load workbook
        print(f"\nReading file: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        print(f"Sheet name: {ws.title}")
        print(f"Total rows: {ws.max_row}")
        print(f"Total columns: {ws.max_column}")
        
        # Find where "IN USD Mn" header is located
        usd_header_col = None
        header_row = None
        
        # Search first 10 rows for "USD" or "IN USD"
        for row_idx in range(1, min(11, ws.max_row + 1)):
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value and "USD" in str(cell_value).upper():
                    usd_header_col = col_idx
                    header_row = row_idx
                    print(f"\n✓ Found 'USD' header at Row {row_idx}, Column {col_idx}")
                    print(f"  Header text: '{cell_value}'")
                    break
            if usd_header_col:
                break
        
        # If no explicit USD header found, assume USD starts from middle
        if not usd_header_col:
            usd_header_col = (ws.max_column // 2) + 1
            print(f"\n⚠ No 'USD' header found, assuming USD starts at column {usd_header_col}")
        
        # Find the sector column (usually column 2 with "Sectors")
        sector_col = 2
        for col_idx in range(1, 5):
            cell_value = ws.cell(row=header_row if header_row else 4, column=col_idx).value
            if cell_value and "Sector" in str(cell_value):
                sector_col = col_idx
                print(f"✓ Found 'Sectors' column at Column {col_idx}")
                break
        
        # Create new workbook for USD data only
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.title = "USD Data Only"
        
        # Determine how many header rows to copy (usually first 3-4 rows)
        header_rows = 4
        
        print(f"\nExtracting data...")
        print(f"  Header rows: {header_rows}")
        print(f"  Sector column: {sector_col}")
        print(f"  USD starts from column: {usd_header_col}")
        
        # Copy all rows
        new_row = 1
        for row_idx in range(1, ws.max_row + 1):
            new_col = 1
            
            # First, copy the sector column (usually Sr. No. and Sectors)
            for col_offset in range(2):  # Copy first 2 columns (Sr. No. and Sectors)
                if sector_col - 1 + col_offset > 0:
                    source_cell = ws.cell(row=row_idx, column=sector_col - 1 + col_offset)
                    target_cell = new_ws.cell(row=new_row, column=new_col)
                    target_cell.value = source_cell.value
                    
                    # Copy formatting
                    if source_cell.font:
                        target_cell.font = Font(
                            bold=source_cell.font.bold,
                            size=source_cell.font.size
                        )
                    if source_cell.fill:
                        target_cell.fill = source_cell.fill
                    if source_cell.alignment:
                        target_cell.alignment = source_cell.alignment
                    
                    new_col += 1
            
            # Then copy all USD columns (from usd_header_col onwards)
            for col_idx in range(usd_header_col, ws.max_column + 1):
                source_cell = ws.cell(row=row_idx, column=col_idx)
                target_cell = new_ws.cell(row=new_row, column=new_col)
                target_cell.value = source_cell.value
                
                # Copy formatting
                if source_cell.font:
                    target_cell.font = Font(
                        bold=source_cell.font.bold,
                        size=source_cell.font.size
                    )
                if source_cell.fill:
                    target_cell.fill = source_cell.fill
                if source_cell.alignment:
                    target_cell.alignment = source_cell.alignment
                
                new_col += 1
            
            new_row += 1
        
        # Apply header formatting
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        for row in range(1, header_rows + 1):
            for col in range(1, new_ws.max_column + 1):
                cell = new_ws.cell(row=row, column=col)
                cell.font = Font(bold=True, size=10)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Auto-adjust column widths
        for col_idx in range(1, new_ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in new_ws[column_letter]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            new_ws.column_dimensions[column_letter].width = max(adjusted_width, 12)
        
        # Save new Excel file
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"AUC_USD_Only_{timestamp}.xlsx"
        new_wb.save(output_file)
        
        print(f"\n{'='*70}")
        print(f"✓ SUCCESS! USD data extracted to: {output_file}")
        print(f"{'='*70}")
        print(f"  Total rows: {new_row - 1}")
        print(f"  Total columns: {new_ws.max_column}")
        print(f"  Header rows preserved: {header_rows}")
        
        # Display preview
        print(f"\nPreview of extracted data:")
        print("-" * 70)
        
        # Show first 10 data rows
        for row_idx in range(1, min(15, new_ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(6, new_ws.max_column + 1)):
                value = new_ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(value)[:20] if value else "")
            print("  ".join(f"{v:20}" for v in row_data))
        
        print("-" * 70)
        
        return output_file
        
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def extract_usd_pandas_method(excel_file):
    """
    Alternative method using pandas - simpler but may lose some formatting
    """
    print("\n" + "=" * 70)
    print("Alternative Method: Pandas-based Extraction")
    print("=" * 70)
    
    try:
        # Read with multiple header rows
        df = pd.read_excel(excel_file, header=[0, 1, 2, 3])
        
        print(f"\nDataFrame shape: {df.shape}")
        print(f"Column levels: {df.columns.nlevels}")
        
        # Find USD columns
        usd_cols = []
        sector_cols = []
        
        for col in df.columns:
            col_str = str(col).upper()
            if 'USD' in col_str or 'Unnamed' not in col_str:
                # Check if it's a sector identifier column or USD column
                if 'SECTOR' in col_str or 'SR' in col_str:
                    sector_cols.append(col)
                elif any(x in col_str for x in ['USD', 'EQUITY', 'DEBT', 'HYBRID']):
                    usd_cols.append(col)
        
        print(f"\nFound {len(sector_cols)} sector columns")
        print(f"Found {len(usd_cols)} USD columns")
        
        # Extract relevant columns
        if len(usd_cols) > 0:
            result_df = df[sector_cols + usd_cols].copy()
            
            output_file = "AUC_USD_Pandas_Extract.xlsx"
            result_df.to_excel(output_file, index=False)
            
            print(f"\n✓ Saved to: {output_file}")
            print(f"\nPreview:")
            print(result_df.head(10))
            
            return output_file
        else:
            print("❌ No USD columns identified")
            return None
            
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    import glob
    import os
    
    # Auto-find the latest NSDL Excel file
    files = glob.glob("NSDL_FII_*.xlsx")
    
    if not files:
        print("❌ No NSDL Excel files found in current directory")
        print("\nPlease ensure you have downloaded the NSDL data first!")
        print("Expected filename pattern: NSDL_FII_*.xlsx")
        exit(1)
    
    # Get most recent file
    excel_file = max(files, key=os.path.getctime)
    
    print("\n")
    print(f"Found file: {excel_file}")
    print(f"File size: {os.path.getsize(excel_file)} bytes")
    
    # Try openpyxl method (preserves formatting)
    print("\n" + "="*70)
    print("Method 1: Using openpyxl (preserves formatting)")
    print("="*70)
    
    result = extract_usd_data_with_headers(excel_file)
    
    if not result:
        # Try pandas method as fallback
        print("\nTrying alternative pandas method...")
        result = extract_usd_pandas_method(excel_file)
    
    if result:
        print("\n" + "="*70)
        print("✓ EXTRACTION COMPLETED SUCCESSFULLY!")
        print("="*70)
        print(f"\nOutput file: {result}")
        print("\nThe extracted file contains:")
        print("  ✓ All header rows including 'AUC as on [Date]'")
        print("  ✓ IN USD Mn heading preserved")
        print("  ✓ Only USD columns (no INR data)")
        print("  ✓ Sector names and Sr. No.")
    else:
        print("\n❌ Extraction failed")
    
    print("\n")